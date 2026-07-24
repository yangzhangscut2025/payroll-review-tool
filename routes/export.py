import os
import json
import uuid
import threading
from datetime import datetime
from flask import Blueprint, request, session, send_file, redirect, url_for, current_app, jsonify
from models import db, Project, ReviewField
from excel_writer import generate_review_excel, LANG_MAP

export_bp = Blueprint('export', __name__)

# In-memory progress store (cleared on restart)
_export_progress = {}

TASK_TTL = 3600  # auto-clean tasks older than 1 hour


def _cleanup_tasks():
    """Remove expired tasks from progress store."""
    import time
    now = time.time()
    expired = [tid for tid, info in _export_progress.items()
               if info.get('_created', 0) < now - TASK_TTL]
    for tid in expired:
        _export_progress.pop(tid, None)


def _find_active_task(project_id):
    """Check if there's already an active export for this project."""
    for tid, info in _export_progress.items():
        if info.get('project_id') == project_id and info['status'] == 'processing':
            return tid
    return None


@export_bp.route('/api/projects/<int:project_id>/export-preview', methods=['POST'])
def export_preview(project_id):
    """Return export preview: sheets, fields, translation estimates."""
    username = session.get('username')
    if not username:
        return jsonify({'error': '未登录'}), 401

    project = Project.query.get_or_404(project_id)
    if not project.file_path:
        return jsonify({'error': '请先上传Excel文件'}), 400

    fields = ReviewField.query.filter_by(project_id=project.id).all()
    if not fields:
        return jsonify({'error': '请先进行审阅'}), 400

    review_map = {}
    for f in fields:
        review_map[(f.row_index, f.field_type)] = f

    # Check which sheets exist
    import openpyxl
    wb = openpyxl.load_workbook(project.file_path, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        if name == 'zh_Chinese' or name in LANG_MAP:
            sheets.append({
                'name': name,
                'label': '中文' if name == 'zh_Chinese' else LANG_MAP.get(name, name),
                'translate': name in LANG_MAP,
            })
    wb.close()

    # Count translatable items (only fields with actual review marks)
    import re
    _mark_pat = re.compile(r'<(span style|/s>|s>|strong>|/strong>)')
    def _has_marks(t):
        return bool(_mark_pat.search(t)) if t else False
    translate_items = sum(1 for rf in review_map.values()
                          if rf.changed_content and _has_marks(rf.changed_content))

    # Estimate tokens and cost (changed fields + English reference)
    est_chars = sum(len(rf.changed_content) for rf in review_map.values()
                    if rf.changed_content and _has_marks(rf.changed_content))
    # Add English reference text (same fields, from English sheet)
    en_chars = 0
    if any(s['translate'] for s in sheets):
        import openpyxl
        en_wb = openpyxl.load_workbook(project.file_path, data_only=True)
        if 'en_English' in en_wb.sheetnames:
            en_ws = en_wb['en_English']
            en_headers = [str(en_ws.cell(row=1, column=c).value or '').strip() for c in range(1, en_ws.max_column + 1)]
            review_cols = ['实务内容（官方）', '实务内容（行业通用）', '参考依据（官方）', '参考依据（行业权威）',
                           '官方规则', '行业通用', '官方网站', '权威网站']
            en_col_map = {h: i+1 for i, h in enumerate(en_headers) if h in review_cols}
            for rf in review_map.values():
                if rf.changed_content and _has_marks(rf.changed_content):
                    col = en_col_map.get(rf.field_type)
                    if col:
                        val = en_ws.cell(row=rf.row_index, column=col).value
                        if val:
                            en_chars += len(str(val))
        en_wb.close()
    est_chars += en_chars
    est_tokens = int(est_chars * 1.5)  # rough: 1.5 tokens per Chinese char
    translate_sheets = sum(1 for s in sheets if s['translate'])
    total_tokens = est_tokens * translate_sheets
    est_cost = (total_tokens / 1_000_000) * 1.0  # input ¥1/M tokens
    est_time = max(5, translate_items * translate_sheets * 0.3)  # ~0.3s per item

    return jsonify({
        'ok': True,
        'sheets': sheets,
        'total_fields': len(fields),
        'translate_items': translate_items,
        'est_tokens': total_tokens,
        'est_cost': round(est_cost, 4),
        'est_time_seconds': round(est_time, 1),
    })


@export_bp.route('/api/projects/<int:project_id>/export-start', methods=['POST'])
def export_start(project_id):
    """Start async export, return task_id for progress polling."""
    username = session.get('username')
    if not username:
        return jsonify({'error': '未登录'}), 401

    project = Project.query.get_or_404(project_id)
    if not project.file_path:
        return jsonify({'error': '请先上传Excel文件'}), 400

    fields = ReviewField.query.filter_by(project_id=project.id).all()
    if not fields:
        return jsonify({'error': '请先进行审阅'}), 400

    review_map = {}
    for f in fields:
        review_map[(f.row_index, f.field_type)] = f

    safe_name = project.name.replace('/', '_').replace('\\', '_')
    date_str = datetime.now().strftime('%Y%m%d')
    output_name = f"{safe_name}_审阅结果_{date_str}.xlsx"
    output_path = os.path.join(os.path.dirname(project.file_path), output_name)

    task_id = uuid.uuid4().hex[:12]

    # Check for existing active export
    existing = _find_active_task(project_id)
    if existing:
        return jsonify({'ok': True, 'task_id': existing, 'resumed': True})

    _cleanup_tasks()
    _export_progress[task_id] = {
        'status': 'processing',
        'percentage': 0,
        'sheet': '',
        'detail': '准备中...',
        'output_path': output_path,
        'output_name': output_name,
        'error': None,
        'project_id': project_id,
        '_created': __import__('time').time(),
    }

    def progress_callback(pct, sheet, detail):
        if task_id in _export_progress:
            _export_progress[task_id].update({
                'percentage': pct,
                'sheet': sheet,
                'detail': detail,
            })

    app = current_app._get_current_object()

    def run_export():
        try:
            warnings, usage = generate_review_excel(
                project.file_path, output_path, review_map,
                project.format_version or 'v1',
                progress_callback=progress_callback,
            )
            _export_progress[task_id]['status'] = 'done'
            _export_progress[task_id]['percentage'] = 100
            detail = '导出完成'
            if warnings:
                detail += ' (⚠️ ' + '; '.join(warnings) + ')'
            _export_progress[task_id]['detail'] = detail
            with app.app_context():
                from models import db, Project, AIUsageLog
                p = db.session.get(Project, project_id)
                if p:
                    p.output_path = output_path
                    p.updated_at = db.func.now()
                # Log translation usage
                if usage['prompt_tokens'] > 0 or usage['completion_tokens'] > 0:
                    log = AIUsageLog(
                        username=username,
                        project_id=project_id,
                        project_name=project.name,
                        module_l1='export',
                        module_l2='translation',
                        model=current_app.config.get('DEEPSEEK_MODEL', 'deepseek-chat'),
                        prompt_tokens=usage['prompt_tokens'],
                        completion_tokens=usage['completion_tokens'],
                    )
                    db.session.add(log)
                db.session.commit()
        except Exception as e:
            _export_progress[task_id]['status'] = 'error'
            _export_progress[task_id]['error'] = str(e)
            _export_progress[task_id]['detail'] = '导出失败'

    thread = threading.Thread(target=run_export, daemon=True)
    thread.start()

    return jsonify({'ok': True, 'task_id': task_id})


@export_bp.route('/api/projects/<int:project_id>/export-progress/<task_id>')
def export_progress(project_id, task_id):
    """Poll for export progress."""
    username = session.get('username')
    if not username:
        return jsonify({'error': '未登录'}), 401

    info = _export_progress.get(task_id)
    if not info:
        return jsonify({'error': '任务不存在或已过期'}), 404

    _cleanup_tasks()

    return jsonify({
        'status': info['status'],
        'percentage': info['percentage'],
        'sheet': info['sheet'],
        'detail': info['detail'],
        'error': info.get('error'),
    })


@export_bp.route('/api/projects/<int:project_id>/export-download/<task_id>')
def export_download(project_id, task_id):
    """Download the exported file."""
    username = session.get('username')
    if not username:
        return redirect(url_for('auth.login_page'))

    info = _export_progress.get(task_id)
    if not info or info['status'] != 'done':
        return "导出尚未完成或已过期", 404

    output_path = info['output_path']
    output_name = info['output_name']

    # Clean up progress
    _export_progress.pop(task_id, None)

    return send_file(output_path, as_attachment=True, download_name=output_name)


@export_bp.route('/projects/<int:project_id>/export')
def export_quick(project_id):
    """Quick export: direct download, no translation, Chinese sheet only."""
    username = session.get('username')
    if not username:
        return redirect(url_for('auth.login_page'))

    project = Project.query.get_or_404(project_id)
    if not project.file_path:
        return "请先上传Excel文件", 400

    fields = ReviewField.query.filter_by(project_id=project.id).all()
    if not fields:
        return "请先进行审阅", 400

    review_map = {}
    for f in fields:
        review_map[(f.row_index, f.field_type)] = f

    safe_name = project.name.replace('/', '_').replace('\\', '_')
    date_str = datetime.now().strftime('%Y%m%d')
    output_name = f"{safe_name}_快速导出_{date_str}.xlsx"
    output_path = os.path.join(os.path.dirname(project.file_path), output_name)

    import openpyxl
    from copy import copy
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    from excel_writer import REVIEW_COLUMNS, _apply_html_to_cell

    review_cols = REVIEW_COLUMNS.get(project.format_version or 'v1', REVIEW_COLUMNS['v1'])
    wb = openpyxl.load_workbook(project.file_path, data_only=True)

    for ws in wb.worksheets:
        if ws.title != 'zh_Chinese':
            continue
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val).strip() if val else '')

        col_positions = {}
        for h in headers:
            if h in review_cols:
                col_positions[h] = headers.index(h) + 1

        sorted_cols = sorted(col_positions.items(), key=lambda x: x[1], reverse=True)

        for field_type, base_col in sorted_cols:
            insert_col = base_col + 1
            ws.insert_cols(insert_col)
            header_cell = ws.cell(row=1, column=insert_col)
            header_cell.value = f"{field_type}_校验列"
            orig_header = ws.cell(row=1, column=base_col)
            header_cell.font = copy(orig_header.font)
            header_cell.alignment = copy(orig_header.alignment)

            for row in range(2, ws.max_row + 1):
                key = (row, field_type)
                if key in review_map:
                    rf = review_map[key]
                    cell = ws.cell(row=row, column=insert_col)
                    bg = {'待审阅': 'F0F0F0', '已确认': '36cf50', '需修改': 'ffb700', '待讨论': '722ed1'}
                    if rf.status != '待审阅':
                        cell.fill = PatternFill(start_color=bg.get(rf.status, 'F0F0F0'),
                                                end_color=bg.get(rf.status, 'F0F0F0'), fill_type='solid')
                    content = rf.changed_content
                    if content:
                        # Format invalid links for reference fields
                        if '参考' in field_type or '网站' in field_type:
                            link_statuses = rf.get_link_statuses()
                            corrected = rf.get_corrected_links()
                            if link_statuses:
                                notes = []
                                for url, status in link_statuses.items():
                                    if status in ('打不开', '内容不符', '已过时'):
                                        marked = f'<span style="color:#C00000;"><s>{url}</s></span>'
                                        if url in corrected:
                                            marked += f' → {corrected[url]}'
                                        if url in content:
                                            content = content.replace(url, marked)
                                        notes.append(f'[{status}] {url}' + (f' → {corrected[url]}' if url in corrected else ''))
                                if notes:
                                    content += '\n\n【链接状态】\n' + '\n'.join(notes)
                        if '<' in content:
                            _apply_html_to_cell(cell, content)
                        else:
                            cell.value = content

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(output_path)
    wb.close()

    return send_file(output_path, as_attachment=True, download_name=output_name)