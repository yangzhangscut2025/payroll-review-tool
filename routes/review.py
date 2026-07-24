import json
import re
from flask import Blueprint, render_template, request, session, jsonify, redirect, url_for
from models import db, Project, ReviewField, ModuleAssignment
from users import get_all_users

review_bp = Blueprint('review', __name__)

FIELD_TYPE_MAP = {
    '实务内容（官方）': '官方',
    '实务内容（行业通用）': '行业',
    '实务内容（内部口径）': '内部',
    '参考依据（官方）': '官方链接',
    '参考依据（行业权威）': '行业权威链接',
    '参考依据（行业常规）': '行业常规链接',
}

PAIRS = {
    'v1': [
        {'content': '实务内容（官方）', 'ref': '参考依据（官方）', 'label': '官方'},
        {'content': '实务内容（行业通用）', 'ref': '参考依据（行业权威）', 'label': '行业通用'},
        {'content': '实务内容（内部口径）', 'ref': '参考依据（行业常规）', 'label': '内部口径'},
    ],
    'v2': [
        {'content': '官方规则', 'ref': '官方网站', 'label': '官方规则'},
        {'content': '行业通用', 'ref': '权威网站', 'label': '行业通用'},
    ],
    'v3': [
        {'content': '实务内容（官方）', 'ref': '参考依据（官方）', 'label': '官方'},
        {'content': '实务内容（行业通用）', 'ref': '参考依据（行业权威）', 'label': '行业通用'},
    ],
}


def url_extract(text):
    """Extract URLs from text, deduplicate."""
    if not text:
        return []
    urls = re.findall(r'https?://[^\s<>"\')\]]+', text)
    seen = set()
    unique = []
    for u in urls:
        # Clean trailing punctuation
        u = u.rstrip('.,;:!?')
        if u not in seen:
            seen.add(u)
            unique.append(u)
    return unique


def check_permission(project, module_l1, username):
    """Check if user can edit this module."""
    if project.created_by == username:
        return True
    assignment = ModuleAssignment.query.filter_by(
        project_id=project.id, module_l1=module_l1).first()
    return assignment and assignment.assignee == username


@review_bp.route('/projects/<int:project_id>/review')
def review_workspace(project_id):
    username = session.get('username')
    if not username:
        return redirect(url_for('auth.login_page'))

    project = Project.query.get_or_404(project_id)
    if not project.file_path:
        return "请先上传Excel文件", 400

    # Get ordered l2 modules
    fields = ReviewField.query.filter_by(project_id=project.id)\
        .order_by(ReviewField.row_index).all()

    l2_list = []
    seen = set()
    for f in fields:
        key = (f.module_l1, f.module_l2)
        if key not in seen:
            seen.add(key)
            l2_list.append({'l1': f.module_l1, 'l2': f.module_l2})

    # Get assignments
    assignments = ModuleAssignment.query.filter_by(project_id=project.id).all()
    assigned_map = {a.module_l1: a for a in assignments}

    # Default to first l2
    l2_index = request.args.get('l2', 0, type=int)

    # For non-owners: auto-redirect to first assigned module
    if project.created_by != username and not request.args.get('l2'):
        for i, item in enumerate(l2_list):
            ma = assigned_map.get(item['l1'])
            if ma and ma.assignee == username:
                l2_index = i
                break

    if l2_index < 0 or l2_index >= len(l2_list):
        l2_index = 0

    current = l2_list[l2_index]
    current_l1 = current['l1']
    current_l2 = current['l2']

    # Get all 6 fields for this l2
    l2_fields = ReviewField.query.filter_by(
        project_id=project.id, module_l1=current_l1, module_l2=current_l2
    ).order_by(ReviewField.id).all()

    # Build pairs
    pairs = PAIRS.get(project.format_version or 'v1', PAIRS['v1'])
    pairs_data = []
    for pair in pairs:
        content_field = None
        ref_field = None
        for f in l2_fields:
            if f.field_type == pair['content']:
                content_field = f
            elif f.field_type == pair['ref']:
                ref_field = f

        if content_field:
            urls = content_field.get_link_urls()
            content_field.link_urls_list = urls if urls else url_extract(content_field.original_content)
        if ref_field:
            urls = ref_field.get_link_urls()
            ref_field.link_urls_list = urls if urls else url_extract(ref_field.original_content)
            ref_field.link_statuses_dict = ref_field.get_link_statuses()
            ref_field.corrected_links_dict = ref_field.get_corrected_links()

        pairs_data.append({
            'label': pair['label'],
            'content': content_field,
            'reference': ref_field,
        })

    # Build tree data with pre-computed l2 indices
    l2_index_map = {}
    for i, item in enumerate(l2_list):
        key = f"{item['l1']}|{item['l2']}"
        l2_index_map[key] = i

    l1_tree = {}
    for item in l2_list:
        l1_tree.setdefault(item['l1'], []).append(item['l2'])

    # Get field statuses for tree
    field_statuses = {}
    l2_status = {}  # l1|l2 -> overall status color
    l2_fields_count = {}  # l1|l2 -> {total, completed}
    for f in fields:
        key = f"{f.module_l1}|{f.module_l2}|{f.field_type}"
        field_statuses[key] = f.status
        l2_key = f"{f.module_l1}|{f.module_l2}"
        if l2_key not in l2_fields_count:
            l2_fields_count[l2_key] = {'total': 0, 'completed': 0}
        l2_fields_count[l2_key]['total'] += 1
        if f.status != '待审阅':
            l2_fields_count[l2_key]['completed'] += 1
    for l2_key, counts in l2_fields_count.items():
        if counts['completed'] == 0:
            l2_status[l2_key] = '#F2F3F5'  # 灰色：未开始
        elif counts['completed'] == counts['total']:
            l2_status[l2_key] = '#36CFC9'  # 绿色：全部完成
        else:
            l2_status[l2_key] = '#FFB800'  # 黄色：进行中

    # Read l2_说明 from Excel
    l2_description = ''
    try:
        import openpyxl
        wb = openpyxl.load_workbook(project.file_path, data_only=True)
        ws = wb.active
        fmt = project.format_version or 'v1'
        l2_col = 3 if fmt in ('v1', 'v3') else 2
        desc_col = 4 if fmt in ('v1', 'v3') else 3
        for row in range(2, ws.max_row + 1):
            l1_val = str(ws.cell(row=row, column=1).value or '').strip()
            l2_val = str(ws.cell(row=row, column=l2_col).value or '').strip()
            if l1_val == current_l1 and l2_val == current_l2:
                desc = ws.cell(row=row, column=desc_col).value
                if desc:
                    l2_description = str(desc).strip()
                break
        wb.close()
    except Exception:
        pass

    # Progress
    total_fields = len(fields)
    completed = sum(1 for f in fields if f.status != '待审阅')

    return render_template('review.html', project=project, username=username,
                           l2_list=l2_list, l2_index=l2_index,
                           l2_list_json=json.dumps(l2_list, ensure_ascii=False),
                           current_l1=current_l1, current_l2=current_l2,
                           l2_description=l2_description,
                           pairs=pairs_data, l1_tree=l1_tree,
                           field_statuses=field_statuses,
                           l2_status=l2_status,
                           assigned_map=assigned_map,
                           l2_index_map=l2_index_map,
                           user_list=get_all_users(),
                           total_fields=total_fields, completed=completed)


@review_bp.route('/api/review/<int:field_id>/save', methods=['POST'])
def save_field(field_id):
    username = session.get('username')
    if not username:
        return jsonify({'error': '未登录'}), 401

    field = ReviewField.query.get_or_404(field_id)
    project = Project.query.get(field.project_id)

    if not check_permission(project, field.module_l1, username):
        return jsonify({'error': '无权限编辑此模块'}), 403

    data = request.get_json()
    if not data:
        return jsonify({'error': '无效请求'}), 400

    if 'status' in data:
        field.status = data['status']
        field.reviewer = username
        field.reviewed_at = db.func.now()
    if 'change_type' in data:
        field.change_type = data['change_type']
    if 'internal_note' in data:
        field.internal_note = data['internal_note']
    if 'changed_content' in data:
        field.changed_content = data['changed_content']
        # Re-extract URLs from changed content and persist
        field.set_link_urls(url_extract(field.changed_content))
    if 'link_statuses' in data:
        field.set_link_statuses(data['link_statuses'])
    if 'corrected_links' in data:
        field.set_corrected_links(data['corrected_links'])

    from sqlalchemy.exc import OperationalError
    import time
    max_retries = 3
    for attempt in range(max_retries):
        try:
            db.session.commit()
            break
        except OperationalError as e:
            db.session.rollback()
            if attempt < max_retries - 1:
                time.sleep(0.3 * (attempt + 1))
            else:
                return jsonify({'error': '数据库忙，请重试'}), 503

    # Update module completion
    _update_module_progress(project, field.module_l1)

    return jsonify({'ok': True, 'status': field.status})


@review_bp.route('/api/projects/<int:project_id>/fields')
def get_fields_status(project_id):
    username = session.get('username')
    if not username:
        return jsonify({'error': '未登录'}), 401

    fields = ReviewField.query.filter_by(project_id=project_id).all()
    statuses = {}
    for f in fields:
        key = f"{f.module_l1}|{f.module_l2}|{f.field_type}"
        statuses[key] = f.status
    return jsonify(statuses)


def _update_module_progress(project, module_l1):
    """Recalculate project and module completion."""
    fields = ReviewField.query.filter_by(project_id=project.id, module_l1=module_l1).all()
    all_done = all(f.status != '待审阅' for f in fields)

    ma = ModuleAssignment.query.filter_by(project_id=project.id, module_l1=module_l1).first()
    if ma:
        ma.status = '已完成' if all_done else '进行中'

    # Count all l1 modules completion
    l1_set = set()
    all_fields = ReviewField.query.filter_by(project_id=project.id).all()
    for f in all_fields:
        l1_set.add(f.module_l1)

    completed_count = 0
    for l1 in l1_set:
        l1_fields = [f for f in all_fields if f.module_l1 == l1]
        if all(f.status != '待审阅' for f in l1_fields):
            completed_count += 1

    project.completed_modules = completed_count
    project.total_modules = len(l1_set)
    project.status = '已完成' if completed_count == len(l1_set) else '进行中'
    db.session.commit()


@review_bp.route('/api/review/<int:project_id>/ai-check', methods=['POST'])
def ai_check(project_id):
    """Proxy to DeepSeek API for AI-powered compliance checking."""
    username = session.get('username')
    if not username:
        return jsonify({'error': '未登录'}), 401

    project = Project.query.get_or_404(project_id)

    data = request.get_json()
    if not data or not data.get('prompt'):
        return jsonify({'error': '请提供核对内容'}), 400

    prompt = data['prompt']

    from flask import current_app
    api_key = current_app.config.get('DEEPSEEK_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'DeepSeek API Key 未配置，请在环境变量中设置 DEEPSEEK_API_KEY'}), 500

    api_url = current_app.config.get('DEEPSEEK_API_URL',
        'https://api.deepseek.com/v1/chat/completions')
    model = current_app.config.get('DEEPSEEK_MODEL', 'deepseek-chat')

    payload = {
        'model': model,
        'messages': [
            {'role': 'user', 'content': prompt},
        ],
        'temperature': 0.3,
        'max_tokens': 4096,
        'stream': False,
    }

    try:
        import requests
        resp = requests.post(
            api_url,
            json=payload,
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            timeout=120,
        )
        resp.raise_for_status()
        result = resp.json()
        content = result['choices'][0]['message']['content']
        usage = result.get('usage', {})

        # Log AI usage
        try:
            from models import AIUsageLog
            log = AIUsageLog(
                username=username,
                project_id=project.id,
                project_name=project.name,
                module_l1=data.get('module_l1', ''),
                module_l2=data.get('module_l2', ''),
                model=result.get('model', model),
                prompt_tokens=usage.get('prompt_tokens', 0),
                completion_tokens=usage.get('completion_tokens', 0),
            )
            db.session.add(log)
            db.session.commit()
        except Exception as log_err:
            current_app.logger.error(f'Failed to log AI usage: {log_err}')

        return jsonify({
            'ok': True,
            'content': content,
            'model': result.get('model', model),
            'usage': {
                'prompt_tokens': usage.get('prompt_tokens', 0),
                'completion_tokens': usage.get('completion_tokens', 0),
            },
        })
    except requests.exceptions.Timeout:
        current_app.logger.error('DeepSeek API timeout')
        return jsonify({'error': 'AI 响应超时（120秒），请稍后重试'}), 504
    except requests.exceptions.RequestException as e:
        current_app.logger.error(f'DeepSeek API error: {e}')
        # Try to extract the API's own error message
        detail = str(e)
        try:
            if e.response is not None:
                body = e.response.json()
                err = body.get('error', {})
                if isinstance(err, dict):
                    detail = err.get('message_zh', err.get('message', str(e)))
                elif isinstance(err, str):
                    detail = err
        except Exception:
            pass
        return jsonify({'error': f'AI 请求失败：{detail}'}), 502
    except (KeyError, IndexError) as e:
        current_app.logger.error(f'DeepSeek API unexpected response: {e}')
        return jsonify({'error': 'AI 返回数据格式异常'}), 502
