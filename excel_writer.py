import re
import json
import requests
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Review column list and their positions
REVIEW_COLUMNS = {
    'v1': [
        '实务内容（官方）', '实务内容（行业通用）', '实务内容（内部口径）',
        '参考依据（官方）', '参考依据（行业权威）', '参考依据（行业常规）',
    ],
    'v2': [
        '官方规则', '行业通用', '官方网站', '权威网站',
    ],
    'v3': [
        '实务内容（官方）', '实务内容（行业通用）',
        '参考依据（官方）', '参考依据（行业权威）',
    ],
}

STATUS_BG_COLORS = {
    '待审阅': 'F0F0F0',
    '已确认': '36cf50',
    '需修改': 'ffb700',
    '待讨论': '722ed1',
}

GREEN = '548235'
RED = 'C00000'
BLUE = '2E75B5'


def _html_to_cell_parts(html_text):
    """Parse Quill HTML and return list of (text, font_dict) tuples.
    Each font_dict contains openpyxl Font constructor kwargs."""
    if not html_text:
        return [('', {})]

    # Remove <p> and <br> tags (convert to newlines)
    text = html_text.replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n')
    text = re.sub(r'</p>\s*<p[^>]*>', '\n', text)
    text = re.sub(r'</?p[^>]*>', '', text)

    parts = []
    _parse_tags(text, 0, {}, parts)
    return parts if parts else [('', {})]


def _parse_tags(s, offset, current_fmt, parts):
    """Recursively parse HTML string, building (text, font) segments."""
    i = offset
    while i < len(s):
        if s[i] == '<':
            end = s.find('>', i)
            if end == -1:
                # No closing bracket, treat rest as plain text
                parts.append((s[i:], dict(current_fmt)))
                return

            tag_content = s[i + 1:end]
            tag_name = tag_content.split()[0].lower().rstrip('/')

            # Self-closing or void elements
            if tag_content.endswith('/') or tag_name in ('br', 'hr', 'img'):
                if tag_name == 'br':
                    parts.append(('\n', dict(current_fmt)))
                i = end + 1
                continue

            # Closing tag
            if tag_name.startswith('/'):
                return  # Return to parent, restoring parent font

            # Opening tag — build font modifiers
            new_fmt = dict(current_fmt)
            if tag_name in ('strong', 'b'):
                new_fmt['bold'] = True
            elif tag_name in ('em', 'i'):
                new_fmt['italic'] = True
            elif tag_name == 'u':
                new_fmt['underline'] = 'single'
            elif tag_name in ('s', 'strike', 'del'):
                new_fmt['strikethrough'] = True
            elif tag_name == 'a':
                new_fmt['underline'] = 'single'
                href_match = re.search(r'href="([^"]*)"', tag_content)
                if href_match:
                    new_fmt['color'] = '0563C1'

            # Inline style attributes
            style_match = re.search(r'style="([^"]*)"', tag_content)
            if style_match:
                style = style_match.group(1)
                if 'color' in style:
                    # Support both hex (#FF0000) and rgb(r, g, b) formats
                    hex_match = re.search(r'color:\s*(#[0-9a-fA-F]{6})', style)
                    rgb_match = re.search(r'color:\s*rgb\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', style)
                    if hex_match:
                        new_fmt['color'] = hex_match.group(1).lstrip('#')
                    elif rgb_match:
                        r, g, b = int(rgb_match.group(1)), int(rgb_match.group(2)), int(rgb_match.group(3))
                        new_fmt['color'] = f'{r:02X}{g:02X}{b:02X}'

            # Recursively parse inner content with new format
            closing = '</' + tag_name + '>'
            close_pos = s.find(closing, end + 1)
            if close_pos == -1:
                # No closing tag, treat inner as plain
                inner_text = s[end + 1:]
                if inner_text:
                    parts.append((inner_text, dict(new_fmt)))
                return
            else:
                _parse_tags(s, end + 1, new_fmt, parts)
                i = close_pos + len(closing)
        else:
            # Plain text
            next_tag = s.find('<', i)
            if next_tag == -1:
                parts.append((s[i:], dict(current_fmt)))
                return
            text_seg = s[i:next_tag]
            if text_seg:
                parts.append((text_seg, dict(current_fmt)))
            i = next_tag


BATCH_SIZE = 20  # items per API call

LANG_MAP = {
    'en_English': '英文',
}

TRANSLATION_PROMPT = (
    '你是薪酬合规领域的专业翻译。请将以下中文审阅批注翻译为{target_lang}。'
    '每条中文对应一条英文原文参考，请参考英文原文的术语和风格来确保翻译准确。'
    '规则：\n'
    '1. 法律/薪酬/税务术语必须与英文原文参考保持一致\n'
    '2. 文本中的 {{__T0__}} {{__T1__}} 等占位符必须原样保留，不能修改或删除\n'
    '3. 编号、数字、百分比、日期、法律条文编号保持原格式\n'
    '4. 输出纯 JSON 数组，每条严格对应输入的一条，不要任何解释文字\n'
    '中文审阅：\n{items}\n英文原文参考：\n{refs}\n输出：'
)

TAG_PATTERN = re.compile(r'(<[^>]+>)')
REVIEW_MARK_PATTERN = re.compile(r'<(span style|/s>|s>|strong>|/strong>)')


def _strip_html(text):
    """Remove HTML tags, return plain text."""
    return TAG_PATTERN.sub('', text) if text else ''


def _has_review_marks(text):
    """Check if text has actual review formatting (not just <p>/<br> wrappers)."""
    return bool(REVIEW_MARK_PATTERN.search(text)) if text else False


def _protect_html(text):
    """Replace HTML tags with placeholders {__Tn__}, return (protected_text, tag_map).
    Tags are sorted by length (longest first) to avoid partial replacement of nested tags."""
    tags = TAG_PATTERN.findall(text)
    tag_map = {}
    for i, tag in enumerate(tags):
        placeholder = f'{{__T{i}__}}'
        tag_map[placeholder] = tag
    # Replace longest placeholders first to avoid partial matches
    for placeholder, tag in sorted(tag_map.items(), key=lambda x: -len(x[1])):
        text = text.replace(tag, placeholder, 1)
    return text, tag_map


def _restore_html(text, tag_map):
    """Restore HTML tags from placeholders."""
    for placeholder, tag in tag_map.items():
        text = text.replace(placeholder, tag)
    return text


def _translate_batch(items, ref_texts, target_lang, api_key, api_url, model, progress_callback=None):
    """Translate a list of Chinese texts to target language using DeepSeek API.
    ref_texts: English original texts for reference (same length as items).
    HTML tags are protected with placeholders during translation.
    Returns (translated_list, total_usage_dict)."""
    total_usage = {'prompt_tokens': 0, 'completion_tokens': 0}
    if not items or not api_key:
        return items, total_usage

    # Protect HTML tags in each item
    tag_maps = []
    protected_items = []
    for text in items:
        if '<' in text:
            protected, tag_map = _protect_html(text)
            protected_items.append(protected)
            tag_maps.append(tag_map)
        else:
            protected_items.append(text)
            tag_maps.append({})

    # Split into batches
    batches = [protected_items[i:i + BATCH_SIZE] for i in range(0, len(protected_items), BATCH_SIZE)]
    ref_batches = [ref_texts[i:i + BATCH_SIZE] for i in range(0, len(ref_texts), BATCH_SIZE)]
    total_batches = len(batches)
    translated = []

    for batch_idx, batch in enumerate(batches):
        if progress_callback:
            progress_callback(batch_idx + 1, total_batches)

        # Retry loop: if AI returns fewer items, retry the missing ones
        pending = list(batch)
        ref_pending = list(ref_batches[batch_idx])
        batch_result = []
        max_retries = 3

        for retry in range(max_retries):
            if not pending:
                break
            items_json = json.dumps(pending, ensure_ascii=False)
            refs_json = json.dumps(ref_pending, ensure_ascii=False)
            prompt = TRANSLATION_PROMPT.format(
                target_lang=target_lang,
                items=items_json,
                refs=refs_json
            )
            try:
                resp = requests.post(
                    api_url,
                    json={
                        'model': model,
                        'messages': [{'role': 'user', 'content': prompt}],
                        'temperature': 0.1,
                        'max_tokens': 4096,
                    },
                    headers={
                        'Authorization': f'Bearer {api_key}',
                        'Content-Type': 'application/json',
                    },
                    timeout=120,
                )
                resp.raise_for_status()
                result = resp.json()
                content = result['choices'][0]['message']['content'].strip()
                total_usage['prompt_tokens'] += result.get('usage', {}).get('prompt_tokens', 0)
                total_usage['completion_tokens'] += result.get('usage', {}).get('completion_tokens', 0)

                if content.startswith('```'):
                    content = content.split('\n', 1)[1]
                    if content.endswith('```'):
                        content = content.rsplit('\n', 1)[0]
                parsed = json.loads(content)

                if isinstance(parsed, list):
                    n = min(len(parsed), len(pending))
                    batch_result.extend(parsed[:n])
                    pending = pending[n:]
                    ref_pending = ref_pending[n:]  # retry the rest
                    if not pending:
                        break
                else:
                    break  # not a list, can't retry
            except Exception:
                break  # network error, give up

        # Fill remaining with original
        if pending:
            batch_result.extend(pending)
        translated.extend(batch_result)

    # Restore HTML tags
    result = []
    for i, text in enumerate(translated):
        if i < len(tag_maps) and tag_maps[i]:
            result.append(_restore_html(text, tag_maps[i]))
        else:
            result.append(text)
    return result, total_usage


def generate_review_excel(input_path, output_path, review_map, format_version='v1', progress_callback=None):
    """Generate the reviewed Excel file — processes ALL sheets.
    progress_callback(pct, sheet_name, detail) called during processing.
    Returns (warnings_list, total_usage_dict)."""
    warnings = []
    total_usage = {'prompt_tokens': 0, 'completion_tokens': 0}
    _state = {'sheet': ''}

    def _progress(pct, detail):
        if progress_callback:
            progress_callback(pct, _state['sheet'], detail)

    review_cols = REVIEW_COLUMNS.get(format_version, REVIEW_COLUMNS['v1'])
    wb = load_workbook(input_path)

    # Identify which sheets need translation (non-Chinese)
    translate_sheets = [s for s in wb.sheetnames if s in LANG_MAP]

    # Pre-translate: collect all changed_content, translate once per language
    translations = {}  # {lang: {key: translated_text}}
    if translate_sheets:
        from config import Config
        api_key = Config.DEEPSEEK_API_KEY
        api_url = Config.DEEPSEEK_API_URL
        model = Config.DEEPSEEK_MODEL

        if api_key:
            # Only translate fields with actual review marks (color, bold, strikethrough)
            text_entries = []
            for key, rf in review_map.items():
                changed = rf.changed_content
                if changed and changed.strip() and _has_review_marks(changed):
                    text_entries.append((key, changed))
            if text_entries:
                keys, texts = zip(*text_entries)
                texts = list(texts)
                total_items = len(texts)
                ...

                # Read English original content for translation reference
                en_refs = []
                en_ws = wb.worksheets[1] if len(wb.worksheets) > 1 else None
                if en_ws:
                    en_headers = []
                    for col in range(1, en_ws.max_column + 1):
                        v = en_ws.cell(row=1, column=col).value
                        en_headers.append(str(v).strip() if v else '')
                    en_col_map = {}
                    for h in en_headers:
                        if h in review_cols:
                            en_col_map[h] = en_headers.index(h) + 1
                    for key, _ in text_entries:
                        row, field_type = key
                        col = en_col_map.get(field_type)
                        if col:
                            val = en_ws.cell(row=row, column=col).value
                            en_refs.append(str(val).strip() if val else '')
                        else:
                            en_refs.append('')
                else:
                    en_refs = [''] * total_items

                for sheet_name in translate_sheets:
                    _progress(10, f'翻译英文...')
                    translated_texts, usage = _translate_batch(
                        texts, en_refs, '英文', api_key, api_url, model,
                        progress_callback=lambda i, t: _progress(
                            10 + int(80 * (i / t) / len(translate_sheets)),
                            f'翻译英文 {i}/{t}'
                        ) if progress_callback else None
                    )
                    total_usage['prompt_tokens'] += usage['prompt_tokens']
                    total_usage['completion_tokens'] += usage['completion_tokens']
                    if len(translated_texts) == len(texts):
                        translations[sheet_name] = dict(zip(keys, translated_texts))
                    else:
                        matched = min(len(translated_texts), len(texts))
                        partial = dict(zip(keys[:matched], translated_texts[:matched]))
                        translations[sheet_name] = partial
                        warnings.append(f'英文 sheet 翻译 {matched}/{len(texts)} 条，其余保持中文')
            else:
                _progress(10, '无需翻译（无改动内容）')
        else:
            warnings.append('未配置API Key，英文 sheet 保持中文原文')

    # Process sheets by position: first = Chinese, second = English, rest = skip
    target_indices = [0]  # Chinese sheet
    en_sheet_idx = None
    if len(wb.worksheets) > 1:
        target_indices.append(1)  # English sheet
        en_sheet_idx = 1
    translate_sheets = [wb.sheetnames[1]] if en_sheet_idx is not None else []

    for ws_idx, ws in enumerate(wb.worksheets):
        if ws_idx not in target_indices:
            continue

        _state['sheet'] = ws.title
        is_translated = (ws_idx == en_sheet_idx)
        total_targets = len(target_indices)
        current_idx = target_indices.index(ws_idx) + 1
        base_pct = 90 if not translate_sheets else 10 + int(80 / len(translate_sheets)) * len(translate_sheets)
        _progress(base_pct + int((current_idx / total_targets) * (100 - base_pct)),
                  f'写入{ws.title}...')
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val).strip() if val else '')

        col_positions = {}
        for h in headers:
            if h in review_cols:
                col_positions[h] = headers.index(h) + 1

        sorted_cols = sorted(col_positions.items(), key=lambda x: x[1], reverse=True)
        is_translated = ws.title in translations

        for field_type, base_col in sorted_cols:
            insert_col = base_col + 1
            ws.insert_cols(insert_col)

            header_cell = ws.cell(row=1, column=insert_col)
            header_cell.value = f"{field_type}_校验列"
            orig_header = ws.cell(row=1, column=base_col)
            header_cell.font = copy(orig_header.font)
            header_cell.alignment = copy(orig_header.alignment)

            for row in range(2, ws.max_row + 1):
                key = (row, field_type)
                if key in review_map:
                    rf = review_map[key]
                    cell = ws.cell(row=row, column=insert_col)

                    content = rf.changed_content
                    fallback = False
                    if is_translated:
                        has_marks = _has_review_marks(rf.changed_content)
                        if key in translations.get(ws.title, {}):
                            content = translations[ws.title][key]
                        elif has_marks:
                            orig_val = ws.cell(row=row, column=base_col).value
                            if orig_val:
                                content = str(orig_val).strip()
                                fallback = True
                        else:
                            orig_val = ws.cell(row=row, column=base_col).value
                            if orig_val:
                                content = str(orig_val).strip()

                    if content:
                        # Format invalid links for reference fields
                        if '参考' in field_type or '网站' in field_type:
                            link_statuses = rf.get_link_statuses()
                            corrected = rf.get_corrected_links()
                            if link_statuses:
                                notes = []
                                for url, status in link_statuses.items():
                                    if status in ('打不开', '内容不符', '已过时'):
                                        # Mark the URL in content: red + strikethrough
                                        marked = f'<span style="color:#C00000;"><s>{url}</s></span>'
                                        if url in corrected:
                                            marked += f' → {corrected[url]}'
                                        if url in content:
                                            content = content.replace(url, marked)
                                        notes.append(f'[{status}] {url}' + (f' → {corrected[url]}' if url in corrected else ''))
                                if notes:
                                    content += '\n\n【链接状态】\n' + '\n'.join(notes)
                        if '<' in content:
                            _apply_html_to_cell(cell, content)
                        else:
                            cell.value = content

                    # Status color
                    bg_color = STATUS_BG_COLORS.get(rf.status)
                    if bg_color and rf.status != '待审阅':
                        cell.fill = PatternFill(start_color=bg_color,
                                                end_color=bg_color, fill_type='solid')
                    # Fallback marker: yellow on top, always visible
                    if fallback:
                        cell.fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(output_path)
    wb.close()
    return warnings, total_usage


def _apply_html_to_cell(cell, html_text):
    """Apply HTML-formatted text to an openpyxl cell as rich text."""
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    parts = _html_to_cell_parts(html_text)

    # Check if any formatting exists
    has_any_fmt = False
    for text, fmt in parts:
        if fmt:
            has_any_fmt = True
            break

    if not has_any_fmt:
        cell.value = ''.join(p[0] for p in parts)
        return

    # Build rich text
    rich_parts = []
    for text, fmt in parts:
        if not text:
            continue
        if fmt:
            # Map to openpyxl font kwargs
            font_args = {}
            if fmt.get('bold'):
                font_args['b'] = True
            if fmt.get('italic'):
                font_args['i'] = True
            if fmt.get('underline'):
                font_args['u'] = 'single'
            if fmt.get('strikethrough'):
                font_args['strike'] = True
            if fmt.get('color'):
                font_args['color'] = fmt['color']
            if font_args:
                f = InlineFont(**font_args)
                rich_parts.append(TextBlock(f, text))
            else:
                rich_parts.append(text)
        else:
            rich_parts.append(text)

    if rich_parts:
        cell.value = CellRichText(*rich_parts)
