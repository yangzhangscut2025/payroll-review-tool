import openpyxl

# V1 格式（13列，3对）
V1_REQUIRED_COLUMNS = [
    'l1_标题', 'l1_说明', 'l2_标题', 'l2_说明',
    '实务内容（官方）', '实务内容（行业通用）', '实务内容（内部口径）',
    '参考依据（官方）', '参考依据（行业权威）', '参考依据（行业常规）',
    '属性', '状态', '内部计算链路',
]

V1_REVIEW_COLUMNS = [
    '实务内容（官方）', '实务内容（行业通用）', '实务内容（内部口径）',
    '参考依据（官方）', '参考依据（行业权威）', '参考依据（行业常规）',
]

# V2 格式（7列，2对）
V2_REQUIRED_COLUMNS = [
    '标题Ⅰ', '标题Ⅱ', '说明',
    '官方规则', '行业通用',
    '官方网站', '权威网站',
]

V2_REVIEW_COLUMNS = [
    '官方规则', '行业通用', '官方网站', '权威网站',
]

# V3 格式（10列，2对，列名同 v1）
V3_REQUIRED_COLUMNS = [
    'l1_标题', 'l1_说明', 'l2_标题', 'l2_说明',
    '实务内容（官方）', '实务内容（行业通用）',
    '参考依据（官方）', '参考依据（行业权威）',
    '状态', None,
]

V3_REVIEW_COLUMNS = [
    '实务内容（官方）', '实务内容（行业通用）',
    '参考依据（官方）', '参考依据（行业权威）',
]


def detect_format(headers):
    """根据表头自动识别格式版本。"""
    first = headers[0] if headers else ''
    if first == '标题Ⅰ':
        return 'v2'
    # v3: 10列，第5列是"实务内容（官方）"且无"实务内容（内部口径）"
    if len(headers) >= 8 and headers[4] == '实务内容（官方）' and '实务内容（内部口径）' not in headers:
        return 'v3'
    return 'v1'


def get_merged_cell_map(ws):
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        min_col = merge_range.min_col
        min_row = merge_range.min_row
        top_left_val = ws.cell(row=min_row, column=min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged[(row, col)] = top_left_val
    return merged


def parse_excel(filepath, project_id, db):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            val = str(val).strip()
        headers.append(val)

    # 自动识别格式
    fmt = detect_format(headers)
    if fmt == 'v2':
        required = V2_REQUIRED_COLUMNS
        review_cols = V2_REVIEW_COLUMNS
    elif fmt == 'v3':
        required = V3_REQUIRED_COLUMNS
        review_cols = V3_REVIEW_COLUMNS
    else:
        required = V1_REQUIRED_COLUMNS
        review_cols = V1_REVIEW_COLUMNS

    # 校验
    for i, req in enumerate(required):
        if i >= len(headers) or headers[i] != req:
            raise ValueError(
                f"表格结构不符合预期，请检查列名是否完整且完全匹配。"
                f"列 {i+1}：期望「{req}」，实际「{headers[i] if i < len(headers) else '无'}」"
            )

    col_map = {}
    for c_idx, h in enumerate(headers):
        if h in review_cols:
            col_map[h] = c_idx + 1

    merged_map = get_merged_cell_map(ws)

    from models import ReviewField
    l1_set = set()
    l2_set = set()
    total_fields = 0

    for row in range(2, ws.max_row + 1):
        row_has_data = False
        for c in range(1, min(len(headers) + 1, ws.max_column + 1)):
            if ws.cell(row=row, column=c).value is not None:
                row_has_data = True
                break
        if not row_has_data:
            continue

        # V1/V3: l1=col1, l2=col3; V2: l1=col1, l2=col2
        l1_col = 1
        l2_col = 3 if fmt in ('v1', 'v3') else 2

        l1_title = ws.cell(row=row, column=l1_col).value
        if not l1_title:
            l1_title = merged_map.get((row, l1_col), '')
        l1_title = str(l1_title).strip() if l1_title else ''

        l2_title = ws.cell(row=row, column=l2_col).value
        if not l2_title:
            l2_title = merged_map.get((row, l2_col), '')
        l2_title = str(l2_title).strip() if l2_title else ''

        if not l1_title or not l2_title:
            continue

        l1_set.add(l1_title)
        l2_set.add(f"{l1_title}|{l2_title}")

        for field_type in review_cols:
            col = col_map.get(field_type)
            if not col:
                continue
            original = ws.cell(row=row, column=col).value
            original_str = str(original).strip() if original else ''

            rf = ReviewField(
                project_id=project_id,
                row_index=row,
                module_l1=l1_title,
                module_l2=l2_title,
                field_type=field_type,
                original_content=original_str,
                changed_content=original_str,
                status='待审阅',
            )
            db.session.add(rf)
            total_fields += 1

    wb.close()
    return {
        'l1_count': len(l1_set),
        'l2_count': len(l2_set),
        'total_fields': total_fields,
        'format_version': fmt,
    }