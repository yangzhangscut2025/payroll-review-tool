import os
import openpyxl

_USER_FILE = os.path.join(os.path.dirname(__file__), '文件说明.xlsx')


def get_all_users():
    """Read user list from 文件说明.xlsx 用户与权限 sheet.
    Returns list of dicts: [{english, chinese, role}, ...]
    """
    try:
        wb = openpyxl.load_workbook(_USER_FILE, data_only=True)
        if '用户与权限' not in wb.sheetnames:
            wb.close()
            return _fallback_users()
        ws = wb['用户与权限']
        users = []
        for row in range(2, ws.max_row + 1):
            english = ws.cell(row=row, column=2).value
            chinese = ws.cell(row=row, column=3).value
            role = ws.cell(row=row, column=4).value
            if english and role:
                users.append({
                    'english': str(english).strip(),
                    'chinese': str(chinese).strip() if chinese else '',
                    'role': str(role).strip(),
                    'display': f"{english} ({chinese})" if chinese else str(english).strip()
                })
        wb.close()
        return users
    except Exception:
        return _fallback_users()


def _fallback_users():
    return [
        {'english': 'Zora', 'chinese': '张阳', 'role': '负责人', 'display': 'Zora (张阳)'},
        {'english': 'Jack', 'chinese': '刘嘉俊', 'role': '负责人', 'display': 'Jack (刘嘉俊)'},
        {'english': 'Elara', 'chinese': '冯馨平', 'role': '校验人', 'display': 'Elara (冯馨平)'},
        {'english': 'Hailey', 'chinese': '吴敏', 'role': '校验人', 'display': 'Hailey (吴敏)'},
        {'english': 'kevin', 'chinese': '江学文', 'role': '校验人', 'display': 'kevin (江学文)'},
        {'english': 'Leon', 'chinese': '郑泽聪', 'role': '校验人', 'display': 'Leon (郑泽聪)'},
        {'english': 'Mercury', 'chinese': '刘皓文', 'role': '负责人', 'display': 'Mercury (刘皓文)'},
        {'english': 'Touyi', 'chinese': '余炯宜', 'role': '校验人', 'display': 'Touyi (余炯宜)'},
        {'english': 'Yee', 'chinese': '陈易蔓', 'role': '校验人', 'display': 'Yee (陈易蔓)'},
    ]


def get_assignable_users(exclude=None):
    """Return users available for module assignment (校验人 + 负责人)."""
    all_users = get_all_users()
    if exclude:
        return [u for u in all_users if u['english'] != exclude]
    return all_users
